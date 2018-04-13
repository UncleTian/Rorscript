import os
import re
import shutil
import datetime
from openpyxl import Workbook

class Node(object):

    def __init__(self, java_name, type_name):
        self.java_name = java_name
        self.type_name = type_name


def result_is_not_none(matcher):
    return matcher is not None and matcher.groups()


def write_to_sheet(tset):
    book = Workbook()
    sheet = book.active
    # to be done
    for i in range(len(tset)):
        sheet.cell(row= i+1, column=1).value = tset[i].java_name
        sheet.cell(row=i+1, column=2).value = tset[i].type_name

    book.save('test.xlsx')


def find_type(file_path, jlist, tset):
    with open(file_path, 'r') as f:
        all_content = f.readlines()
        java_name = os.path.basename(file_path)

        rex = re.compile(r'\(([\w]+)[\)\s]+Registry[\.\s]+getService')
        for content in all_content:
            result = rex.search(content)
            if result_is_not_none(result):
               # jlist.add(java_name)
                print(java_name + " " + result.group(1))
                node = Node(java_name, result.group(1))
                tset.append(node)
               # tset.add(result.group(1))


def explore(current_dir):
    java_name_list = []
    types_set = []

    for root, _, files in os.walk(current_dir):
        file_list = [file for file in files if ".java" in file]
        for file in file_list:
            file_path = os.path.join(root, os.path.basename(file))
            find_type(file_path=file_path, jlist=java_name_list, tset=types_set)

    write_to_sheet(types_set)



def main():
    current_dir = os.path.abspath(".")
    explore(current_dir)


if __name__ == "__main__":
    start_time = datetime.datetime.now()
    main()
    print("done")
    end_time = datetime.datetime.now()
    total = end_time - start_time

#    minutes, seconds = total // 60, total % 60
    print("used time: " + str(total))
#    print("used time: " + str(minutes) + ":"+ str(seconds).zfill(2))
